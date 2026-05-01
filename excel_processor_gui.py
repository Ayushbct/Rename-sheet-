import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import openpyxl
from openpyxl.utils import column_index_from_string
import os
from pathlib import Path
from threading import Thread

class ExcelProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Sheet Processor")
        self.root.geometry("640x560")
        self.root.resizable(True, True)
        
        self.selected_folder = tk.StringVar()
        self.column1_var = tk.StringVar(value="Z")
        self.column2_var = tk.StringVar(value="AA")
        
        # Title
        title_label = tk.Label(root, text="Excel Sheet Processor", font=("Arial", 16, "bold"))
        title_label.pack(pady=10)
        
        # Folder Selection Frame
        folder_frame = tk.Frame(root)
        folder_frame.pack(pady=10, padx=10, fill=tk.X)
        
        tk.Label(folder_frame, text="Selected Folder:", font=("Arial", 10)).pack(anchor=tk.W)
        
        folder_entry = tk.Entry(folder_frame, textvariable=self.selected_folder, state='readonly', width=70)
        folder_entry.pack(fill=tk.X, pady=5)
        
        # Browse Button
        browse_btn = tk.Button(folder_frame, text="Browse Folder", command=self.browse_folder, bg="#4CAF50", fg="white", padx=10)
        browse_btn.pack(pady=5)
        
        # Column Selection Frame
        column_frame = tk.Frame(root)
        column_frame.pack(pady=10, padx=10, fill=tk.X)
        
        tk.Label(column_frame, text="Column to clean #1:", font=("Arial", 10)).grid(row=0, column=0, sticky=tk.W, padx=(0, 8))
        tk.Label(column_frame, text="Column to clean #2:", font=("Arial", 10)).grid(row=1, column=0, sticky=tk.W, padx=(0, 8))
        
        column1_entry = tk.Entry(column_frame, textvariable=self.column1_var, width=8)
        column1_entry.grid(row=0, column=1, sticky=tk.W)
        
        column2_entry = tk.Entry(column_frame, textvariable=self.column2_var, width=8)
        column2_entry.grid(row=1, column=1, sticky=tk.W)
        
        tk.Label(column_frame, text="Enter letters like Z and AA", font=("Arial", 8), fg="#555555").grid(row=0, column=2, rowspan=2, sticky=tk.W, padx=(16,0))
        
        # Process Button
        process_btn = tk.Button(root, text="Process Excel Files", command=self.start_processing, bg="#2196F3", fg="white", padx=20, pady=10, font=("Arial", 12))
        process_btn.pack(pady=10)
        
        # Output Text Area
        tk.Label(root, text="Processing Log:", font=("Arial", 10, "bold")).pack(anchor=tk.W, padx=10)
        
        self.output_text = scrolledtext.ScrolledText(root, height=15, width=70, state=tk.DISABLED)
        self.output_text.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)
        
        # Status Bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = tk.Label(root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)
    
    def browse_folder(self):
        folder = filedialog.askdirectory(title="Select the folder containing Excel files")
        if folder:
            self.selected_folder.set(folder)
            self.log_message(f"Folder selected: {folder}\n")
    
    def log_message(self, message):
        self.output_text.config(state=tk.NORMAL)
        self.output_text.insert(tk.END, message)
        self.output_text.see(tk.END)
        self.output_text.config(state=tk.DISABLED)
        self.root.update()
    
    def parse_column(self, column_name):
        """Convert a column letter like Z or AA into a 0-based index."""
        if not column_name or not column_name.strip():
            raise ValueError("Column name cannot be empty.")
        try:
            return column_index_from_string(column_name.strip()) - 1
        except Exception:
            raise ValueError(f"Invalid column name: {column_name}")
    
    def clean_columns(self, excel_file_path, col1_name, col2_name):
        """Clean the specified columns by removing spaces and hyphens."""
        try:
            col1_idx = self.parse_column(col1_name)
            col2_idx = self.parse_column(col2_name)
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet = workbook.active
            
            cleaned_count = 0
            
            for row in sheet.iter_rows(min_row=2):
                if col1_idx < len(row):
                    cell1 = row[col1_idx]
                    if cell1.value:
                        original_value = str(cell1.value)
                        cleaned_value = original_value.replace(" ", "").replace("-", "")
                        if cleaned_value != original_value:
                            cell1.value = cleaned_value
                            cleaned_count += 1
                
                if col2_idx < len(row):
                    cell2 = row[col2_idx]
                    if cell2.value:
                        original_value = str(cell2.value)
                        cleaned_value = original_value.replace(" ", "").replace("-", "")
                        if cleaned_value != original_value:
                            cell2.value = cleaned_value
                            cleaned_count += 1
            
            workbook.save(excel_file_path)
            return cleaned_count
            
        except ValueError as e:
            raise
        except Exception as e:
            raise Exception(f"Error cleaning columns: {e}")
    
    def rename_sheet_to_sheet1(self, excel_file_path):
        """Rename sheet to Sheet1 if there's only one sheet"""
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet_names = workbook.sheetnames
            
            if len(sheet_names) > 1:
                return False, f"File has {len(sheet_names)} sheets. Skipped."
            
            if sheet_names[0] != "Sheet1":
                sheet = workbook[sheet_names[0]]
                old_name = sheet_names[0]
                sheet.title = "Sheet1"
                workbook.save(excel_file_path)
                return True, f"Renamed sheet '{old_name}' to 'Sheet1'"
            else:
                return True, "Sheet already named 'Sheet1'"
            
        except Exception as e:
            raise Exception(f"Error renaming sheet: {e}")
    
    def process_excel_files(self, folder_path):
        """Process all Excel files in the folder"""
        try:
            if not os.path.isdir(folder_path):
                self.log_message("❌ Error: Invalid folder path\n")
                return
            
            excel_files = list(Path(folder_path).glob("*.xlsx")) + list(Path(folder_path).glob("*.xls"))
            
            if not excel_files:
                self.log_message("❌ Error: No Excel files found in the selected folder\n")
                self.status_var.set("No files found")
                return
            
            self.log_message(f"Found {len(excel_files)} Excel file(s)\n")
            self.log_message("=" * 60 + "\n")
            
            successful = 0
            failed = 0
            
            col1_name = self.column1_var.get().strip()
            col2_name = self.column2_var.get().strip()
            
            if not col1_name or not col2_name:
                self.log_message("❌ Error: Both column names must be provided.\n")
                self.status_var.set("Invalid columns")
                return
            
            self.log_message(f"Cleaning columns: {col1_name.upper()} and {col2_name.upper()}\n")
            self.log_message("=" * 60 + "\n")
            
            for idx, excel_file in enumerate(excel_files, 1):
                try:
                    filename = os.path.basename(excel_file)
                    self.log_message(f"\n[{idx}/{len(excel_files)}] Processing: {filename}\n")
                    self.status_var.set(f"Processing: {filename}")
                    
                    # Clean columns
                    cleaned_count = self.clean_columns(str(excel_file), col1_name, col2_name)
                    self.log_message(f"  ✓ Cleaned {cleaned_count} cells in columns {col1_name.upper()} & {col2_name.upper()}\n")
                    
                    # Rename sheet
                    success, message = self.rename_sheet_to_sheet1(str(excel_file))
                    self.log_message(f"  ✓ {message}\n")
                    
                    self.log_message(f"✓ Successfully processed: {filename}\n")
                    successful += 1
                    
                except Exception as e:
                    self.log_message(f"✗ Error: {str(e)}\n")
                    failed += 1
            
            self.log_message("=" * 60 + "\n")
            self.log_message(f"\nProcessing Complete!\n")
            self.log_message(f"  ✓ Successful: {successful}\n")
            self.log_message(f"  ✗ Failed: {failed}\n")
            
            self.status_var.set(f"Done! Processed {successful} file(s)")
            messagebox.showinfo("Success", f"Processing complete!\n\nSuccessful: {successful}\nFailed: {failed}")
            
        except Exception as e:
            self.log_message(f"❌ Error: {str(e)}\n")
            self.status_var.set("Error occurred")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
    
    def start_processing(self):
        folder_path = self.selected_folder.get()
        
        if not folder_path:
            messagebox.showwarning("Warning", "Please select a folder first!")
            return
        
        # Clear previous output
        self.output_text.config(state=tk.NORMAL)
        self.output_text.delete(1.0, tk.END)
        self.output_text.config(state=tk.DISABLED)
        
        # Run processing in a separate thread to keep GUI responsive
        thread = Thread(target=self.process_excel_files, args=(folder_path,))
        thread.daemon = True
        thread.start()


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorGUI(root)
    root.mainloop()
