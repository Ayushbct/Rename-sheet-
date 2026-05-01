import openpyxl
import os
import sys
from pathlib import Path

def clean_columns(excel_file_path):
    """
    Clean columns Z (engineno) and AA (chasisno) by removing spaces and hyphens.
    
    Args:
        excel_file_path: Path to the Excel file
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        
        print(f"  Cleaning columns Z (engineno) and AA (chasisno)...")
        
        cleaned_count = 0
        
        # Iterate through rows and clean the columns
        for row in sheet.iter_rows(min_row=2):  # Start from row 2 (skip header)
            # Clean column Z (engineno)
            cell_z = row[25]  # Column Z is index 25 (0-based)
            if cell_z.value:
                original_value = str(cell_z.value)
                cleaned_value = original_value.replace(" ", "").replace("-", "")
                if cleaned_value != original_value:
                    cell_z.value = cleaned_value
                    cleaned_count += 1
            
            # Clean column AA (chasisno)
            cell_aa = row[26]  # Column AA is index 26 (0-based)
            if cell_aa.value:
                original_value = str(cell_aa.value)
                cleaned_value = original_value.replace(" ", "").replace("-", "")
                if cleaned_value != original_value:
                    cell_aa.value = cleaned_value
                    cleaned_count += 1
        
        # Save the modified workbook
        workbook.save(excel_file_path)
        print(f"  Cleaned {cleaned_count} cells")
        
    except Exception as e:
        print(f"  Error cleaning columns: {e}")
        raise

def rename_sheet_to_sheet1(excel_file_path):
    """
    Rename sheet to 'Sheet1' if there is only one sheet and it has a different name.
    Skip if there are multiple sheets.
    
    Args:
        excel_file_path: Path to the Excel file
    """
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file_path)
        
        # Get all sheet names
        sheet_names = workbook.sheetnames
        
        # Check if there are multiple sheets
        if len(sheet_names) > 1:
            print(f"  Warning: File contains {len(sheet_names)} sheets. Skipping rename...")
            print(f"  Sheets found: {', '.join(sheet_names)}")
            return False
        
        # Rename the sheet if it's not "Sheet1"
        if sheet_names[0] != "Sheet1":
            sheet = workbook[sheet_names[0]]
            old_name = sheet_names[0]
            sheet.title = "Sheet1"
            print(f"  Renamed sheet '{old_name}' to 'Sheet1'")
            
            # Save the modified workbook
            workbook.save(excel_file_path)
        else:
            print(f"  Sheet is already named 'Sheet1'.")
        
        return True
        
    except Exception as e:
        print(f"  Error renaming sheet: {e}")
        raise

def process_excel_file(excel_file_path):
    """
    Process a single Excel file: clean columns and rename sheet.
    
    Args:
        excel_file_path: Path to the Excel file
    """
    try:
        print(f"Processing: {os.path.basename(excel_file_path)}")
        
        # Clean columns
        clean_columns(excel_file_path)
        
        # Rename sheet
        rename_sheet_to_sheet1(excel_file_path)
        
        print(f"✓ Successfully processed: {os.path.basename(excel_file_path)}\n")
        return True
        
    except Exception as e:
        print(f"✗ Error processing file: {e}\n")
        return False

def process_all_excel_files(folder_path):
    """
    Process all Excel files in a folder.
    
    Args:
        folder_path: Path to the folder containing Excel files
    """
    # Check if folder exists
    if not os.path.isdir(folder_path):
        print(f"Error: Folder not found - {folder_path}")
        sys.exit(1)
    
    # Find all Excel files
    excel_files = list(Path(folder_path).glob("*.xlsx")) + list(Path(folder_path).glob("*.xls"))
    
    if not excel_files:
        print(f"Error: No Excel files found in {folder_path}")
        sys.exit(1)
    
    print(f"Found {len(excel_files)} Excel file(s) in {folder_path}\n")
    
    # Process each file
    successful = 0
    failed = 0
    
    for excel_file in excel_files:
        if process_excel_file(str(excel_file)):
            successful += 1
        else:
            failed += 1
    
    # Summary
    print("=" * 50)
    print(f"Processing Summary:")
    print(f"  Successful: {successful}")
    print(f"  Failed: {failed}")
    print("=" * 50)

if __name__ == "__main__":
    # Define the folder path containing Excel files
    folder_path = "input_files"
    
    # Process all Excel files in the folder
    process_all_excel_files(folder_path)
